from openpyxl import Workbook,load_workbook
def writeexcel():
    wb = Workbook()
    sheet = wb.active

    sheet['B9'] = 'this is b9'
    sheet['C9'] = '180,18,18'
    sheet['A1'] = 'hello word'
    print(sheet['B9'].value)
    print(sheet['A1'].value)
    sheet.append(['lili','170','40'])
    wb.save('excel_test.xlsx')
    #打开已有文件
    # wb1 =load_workbook('excel_test.xlsx')

def openexcel():
    print('aaaaaa')


if __name__ == '__main__':
    openexcel()